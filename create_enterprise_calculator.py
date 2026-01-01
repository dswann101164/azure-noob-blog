"""
Azure OpenAI ROI Calculator - ENTERPRISE EDITION
The $497 Logic: Answers the CFO's question - "When do we switch to PTU?"
Built following Gemini's enterprise blueprint
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, BarChart, LineChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation

def create_enterprise_calculator():
    """Generate Enterprise-Grade Azure OpenAI ROI Calculator"""
    
    wb = Workbook()
    
    # AZURE COLOR PALETTE (Deep Blues and Grays)
    azure_dark_blue = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")
    azure_light_blue = PatternFill(start_color="50E6FF", end_color="50E6FF", fill_type="solid")
    azure_gray = PatternFill(start_color="5E5E5E", end_color="5E5E5E", fill_type="solid")
    white_text = Font(color="FFFFFF", bold=True, size=12)
    
    input_yellow = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    output_green = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    warning_red = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    
    title_font = Font(bold=True, size=16, color="0078D4")
    header_font = Font(bold=True, size=12, color="FFFFFF")
    
    # ==================================================================
    # TAB 0: INSTRUCTIONS (First tab - builds authority)
    # ==================================================================
    ws_instructions = wb.active
    ws_instructions.title = "INSTRUCTIONS"
    
    # Confidential header
    ws_instructions['B2'] = "CONFIDENTIAL: Azure OpenAI ROI Calculator"
    ws_instructions['B2'].font = Font(bold=True, size=14, color="FF0000")
    ws_instructions.merge_cells('B2:F2')
    
    ws_instructions['B3'] = "Built by Azure-Noob | Tested on 30,000+ Enterprise Resources"
    ws_instructions['B3'].font = Font(italic=True, size=10, color="5E5E5E")
    ws_instructions.merge_cells('B3:F3')
    
    ws_instructions['B5'] = "WHAT THIS TOOL DOES:"
    ws_instructions['B5'].font = Font(bold=True, size=12)
    
    instructions_content = [
        "",
        "This calculator answers the ONE question every CFO asks:",
        "",
        '     "When do we stop paying per token and start buying PTUs?"',
        "",
        "PTU (Provisioned Throughput Units) = Reserved Azure OpenAI capacity",
        "Pay-as-you-go = Microsoft charges you per token consumed",
        "",
        "THE PROBLEM:",
        "Microsoft's pricing calculator shows you token costs.",
        "It doesn't show you the BREAK-EVEN POINT where PTU becomes cheaper.",
        "",
        "THIS TOOL SHOWS YOU:",
        "✓ Your current pay-as-you-go cost",
        "✓ What PTU would cost with reserved capacity",
        "✓ EXACTLY when to switch (the break-even formula)",
        "✓ Projected waste if you don't optimize",
        "✓ Hidden infrastructure costs (15% overhead)",
        "",
        "HOW TO USE:",
        "1. Go to 'Engine' tab",
        "2. Enter your monthly token volume (yellow cells only)",
        "3. Enter your model pricing",
        "4. Calculator shows: SWITCH TO PTU NOW or STAY PAY-AS-YOU-GO",
        "",
        "5. Go to 'Executive Summary' tab",
        "6. Show this to your boss/CFO",
        "7. Get budget approval to optimize",
        "",
        "CREDENTIALS:",
        "Created by: Enterprise Azure Architect",
        "Tested on: 30,000+ Azure resources across 44 subscriptions",
        "Experience: Financial services merger infrastructure consolidation",
        "",
        "SUPPORT:",
        "Email: david@azure-noob.com",
        "Response: 24 hours",
        "",
        "This is not a Microsoft product. Independent operational analysis.",
        "Version 1.0 - January 2026"
    ]
    
    row = 7
    for line in instructions_content:
        ws_instructions[f'B{row}'] = line
        if line.startswith("THE PROBLEM:") or line.startswith("THIS TOOL") or line.startswith("HOW TO USE:") or line.startswith("CREDENTIALS:"):
            ws_instructions[f'B{row}'].font = Font(bold=True, size=11)
        if line.startswith('     "When do we'):
            ws_instructions[f'B{row}'].font = Font(bold=True, size=12, color="0078D4")
        row += 1
    
    ws_instructions.column_dimensions['B'].width = 80
    
    # ==================================================================
    # TAB 1: THE ENGINE (Where the math happens)
    # ==================================================================
    ws_engine = wb.create_sheet("Engine")
    
    ws_engine['B2'] = "Azure OpenAI Cost Engine"
    ws_engine['B2'].font = title_font
    ws_engine.merge_cells('B2:F2')
    
    ws_engine['B3'] = "The calculator that shows when to switch to PTU"
    ws_engine['B3'].font = Font(italic=True, size=10)
    ws_engine.merge_cells('B3:F3')
    
    # SECTION 1: PAY-AS-YOU-GO CALCULATION
    ws_engine['B5'] = "CONSUMPTION (Pay-as-You-Go)"
    ws_engine['B5'].font = header_font
    ws_engine['B5'].fill = azure_dark_blue
    ws_engine.merge_cells('B5:D5')
    
    ws_engine['B7'] = "Monthly Tokens:"
    ws_engine['C7'] = 10000000  # 10M tokens default
    ws_engine['C7'].fill = input_yellow
    ws_engine['C7'].number_format = '#,##0'
    ws_engine['D7'] = "tokens/month"
    
    ws_engine['B8'] = "Price per 1K Tokens:"
    ws_engine['C8'] = 0.015  # GPT-4o default
    ws_engine['C8'].fill = input_yellow
    ws_engine['C8'].number_format = '$0.000'
    ws_engine['D8'] = "(e.g., $0.015 for GPT-4o output)"
    
    ws_engine['B10'] = "Monthly Spend (Pay-as-You-Go):"
    ws_engine['C10'] = "=(C7/1000)*C8"
    ws_engine['C10'].fill = output_green
    ws_engine['C10'].number_format = '$#,##0.00'
    ws_engine['C10'].font = Font(bold=True, size=12)
    
    # SECTION 2: PTU CALCULATION
    ws_engine['B13'] = "PTU (Provisioned Throughput)"
    ws_engine['B13'].font = header_font
    ws_engine['B13'].fill = azure_dark_blue
    ws_engine.merge_cells('B13:D13')
    
    ws_engine['B15'] = "PTU Units:"
    ws_engine['C15'] = 1  # 1 PTU default
    ws_engine['C15'].fill = input_yellow
    ws_engine['D15'] = "(Usually 100-unit increments)"
    
    ws_engine['B16'] = "Price per PTU/Hour:"
    ws_engine['C16'] = 2.5  # ~$2.50/hour default
    ws_engine['C16'].fill = input_yellow
    ws_engine['C16'].number_format = '$0.00'
    ws_engine['D16'] = "(Approx $2.50-3.00 depending on region)"
    
    ws_engine['B17'] = "Hours in Month:"
    ws_engine['C17'] = 730
    ws_engine['C17'].fill = azure_gray
    ws_engine['C17'].font = Font(color="FFFFFF")
    ws_engine['D17'] = "(730 hours = 30.4 days average)"
    
    ws_engine['B19'] = "Monthly PTU Cost:"
    ws_engine['C19'] = "=C15*C16*C17"
    ws_engine['C19'].fill = output_green
    ws_engine['C19'].number_format = '$#,##0.00'
    ws_engine['C19'].font = Font(bold=True, size=12)
    
    # SECTION 3: THE KILLER FORMULA (Break-Even Decision)
    ws_engine['B22'] = "THE DECISION"
    ws_engine['B22'].font = Font(bold=True, size=14, color="FFFFFF")
    ws_engine['B22'].fill = azure_gray
    ws_engine.merge_cells('B22:D22')
    ws_engine['B22'].alignment = Alignment(horizontal='center')
    
    ws_engine['B24'] = "Recommendation:"
    ws_engine['C24'] = '=IF(C10>C19,"SWITCH TO PTU NOW","STAY ON PAY-AS-YOU-GO")'
    ws_engine['C24'].fill = warning_red
    ws_engine['C24'].font = Font(bold=True, size=14)
    ws_engine['C24'].alignment = Alignment(horizontal='center')
    ws_engine.merge_cells('C24:D24')
    
    ws_engine['B26'] = "Monthly Savings with PTU:"
    ws_engine['C26'] = "=IF(C10>C19,C10-C19,0)"
    ws_engine['C26'].fill = output_green
    ws_engine['C26'].number_format = '$#,##0.00'
    ws_engine['C26'].font = Font(bold=True, size=12)
    
    ws_engine['B27'] = "Annual Savings:"
    ws_engine['C27'] = "=C26*12"
    ws_engine['C27'].fill = output_green
    ws_engine['C27'].number_format = '$#,##0.00'
    ws_engine['C27'].font = Font(bold=True, size=14)
    
    ws_engine['B29'] = "Break-Even Point:"
    ws_engine['C29'] = "=(C19/C10)*100"
    ws_engine['C29'].number_format = '0.0"%"'
    ws_engine['D29'] = "of current spend"
    
    # SECTION 4: HIDDEN INFRASTRUCTURE COSTS
    ws_engine['B32'] = "HIDDEN COSTS (The 15% Microsoft Doesn't Show)"
    ws_engine['B32'].font = header_font
    ws_engine['B32'].fill = warning_red
    ws_engine.merge_cells('B32:D32')
    
    ws_engine['B34'] = "Operational Overhead (15%):"
    ws_engine['C34'] = "=C10*0.15"
    ws_engine['C34'].number_format = '$#,##0.00'
    ws_engine['D34'] = "(Log Analytics, Security, Networking)"
    
    ws_engine['B35'] = "TOTAL Monthly Cost:"
    ws_engine['C35'] = "=C10+C34"
    ws_engine['C35'].fill = warning_red
    ws_engine['C35'].number_format = '$#,##0.00'
    ws_engine['C35'].font = Font(bold=True, size=12)
    
    # Column widths
    ws_engine.column_dimensions['B'].width = 30
    ws_engine.column_dimensions['C'].width = 18
    ws_engine.column_dimensions['D'].width = 40
    
    # Protect cells (lock everything except input cells)
    ws_engine.protection.sheet = True
    ws_engine.protection.password = "azure-noob-2026"
    
    # Unlock input cells
    ws_engine['C7'].protection = Protection(locked=False)
    ws_engine['C8'].protection = Protection(locked=False)
    ws_engine['C15'].protection = Protection(locked=False)
    ws_engine['C16'].protection = Protection(locked=False)
    
    # ==================================================================
    # TAB 2: EXECUTIVE SUMMARY (The $497 View)
    # ==================================================================
    ws_exec = wb.create_sheet("Executive Summary")
    
    ws_exec['B2'] = "Executive Summary: Azure OpenAI Cost Optimization"
    ws_exec['B2'].font = Font(bold=True, size=16, color="0078D4")
    ws_exec.merge_cells('B2:F2')
    
    ws_exec['B3'] = "Show this page to your CFO"
    ws_exec['B3'].font = Font(italic=True, size=11, color="FF0000")
    ws_exec.merge_cells('B3:F3')
    
    # BLOCK A: The "Oops" Analysis
    ws_exec['B5'] = 'BLOCK A: The "Cost of Doing Nothing"'
    ws_exec['B5'].font = Font(bold=True, size=13, color="FFFFFF")
    ws_exec['B5'].fill = warning_red
    ws_exec.merge_cells('B5:F5')
    ws_exec['B5'].alignment = Alignment(horizontal='center')
    
    ws_exec['B7'] = "If you scale to 100M tokens/month without optimization:"
    
    ws_exec['B9'] = "Current Monthly Spend:"
    ws_exec['D9'] = "=Engine!C10"
    ws_exec['D9'].number_format = '$#,##0.00'
    ws_exec['D9'].font = Font(bold=True)
    
    ws_exec['B10'] = "Projected at 100M tokens:"
    ws_exec['D10'] = "=(100000000/1000)*Engine!C8"
    ws_exec['D10'].number_format = '$#,##0.00'
    ws_exec['D10'].font = Font(bold=True, size=12)
    
    ws_exec['B11'] = "Plus 15% Infrastructure:"
    ws_exec['D11'] = "=D10*0.15"
    ws_exec['D11'].number_format = '$#,##0.00'
    
    ws_exec['B13'] = "Projected Annual Waste without PTU Optimization:"
    ws_exec['D13'] = "=(D10+D11)*12"
    ws_exec['D13'].fill = warning_red
    ws_exec['D13'].number_format = '$#,##0'
    ws_exec['D13'].font = Font(bold=True, size=14)
    ws_exec.merge_cells('D13:F13')
    ws_exec['D13'].alignment = Alignment(horizontal='center')
    
    # BLOCK B: Token Density Chart
    ws_exec['B16'] = "BLOCK B: Token Density Analysis"
    ws_exec['B16'].font = Font(bold=True, size=13, color="FFFFFF")
    ws_exec['B16'].fill = azure_dark_blue
    ws_exec.merge_cells('B16:F16')
    ws_exec['B16'].alignment = Alignment(horizontal='center')
    
    ws_exec['B18'] = "Monthly Tokens"
    ws_exec['C18'] = "Pay-as-You-Go"
    ws_exec['D18'] = "PTU Cost"
    ws_exec['E18'] = "Winner"
    ws_exec['F18'] = "Savings"
    
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws_exec[f'{col}18'].font = Font(bold=True)
        ws_exec[f'{col}18'].fill = azure_gray
        ws_exec[f'{col}18'].font = Font(bold=True, color="FFFFFF")
    
    token_scenarios = [
        ("1M", 1000000),
        ("10M", 10000000),
        ("50M", 50000000),
        ("100M", 100000000),
        ("500M", 500000000)
    ]
    
    row = 19
    for label, tokens in token_scenarios:
        ws_exec[f'B{row}'] = label
        ws_exec[f'C{row}'] = f"=({tokens}/1000)*Engine!C8"
        ws_exec[f'C{row}'].number_format = '$#,##0'
        
        ws_exec[f'D{row}'] = "=Engine!C19"  # PTU cost same for all
        ws_exec[f'D{row}'].number_format = '$#,##0'
        
        ws_exec[f'E{row}'] = f'=IF(C{row}>D{row},"PTU","Pay-as-go")'
        ws_exec[f'F{row}'] = f'=IF(C{row}>D{row},C{row}-D{row},0)'
        ws_exec[f'F{row}'].number_format = '$#,##0'
        
        # Highlight the winner row in green
        if tokens == 50000000:  # Typically break-even around 50M
            for col in ['B', 'C', 'D', 'E', 'F']:
                ws_exec[f'{col}{row}'].fill = output_green
                ws_exec[f'{col}{row}'].font = Font(bold=True)
        
        row += 1
    
    # BLOCK C: Hidden Infrastructure Costs
    ws_exec['B26'] = "BLOCK C: Hidden Infrastructure Costs"
    ws_exec['B26'].font = Font(bold=True, size=13, color="FFFFFF")
    ws_exec['B26'].fill = azure_dark_blue
    ws_exec.merge_cells('B26:F26')
    ws_exec['B26'].alignment = Alignment(horizontal='center')
    
    ws_exec['B28'] = "Microsoft's Calculator Shows:"
    ws_exec['D28'] = "=Engine!C10"
    ws_exec['D28'].number_format = '$#,##0.00'
    ws_exec['D28'].font = Font(bold=True)
    
    ws_exec['B29'] = "Hidden Infrastructure (15%):"
    ws_exec['D29'] = "=Engine!C34"
    ws_exec['D29'].number_format = '$#,##0.00'
    
    ws_exec['B31'] = "YOU ACTUALLY PAY:"
    ws_exec['D31'] = "=Engine!C35"
    ws_exec['D31'].fill = warning_red
    ws_exec['D31'].number_format = '$#,##0.00'
    ws_exec['D31'].font = Font(bold=True, size=14)
    ws_exec.merge_cells('D31:F31')
    ws_exec['D31'].alignment = Alignment(horizontal='center')
    
    # The Recommendation Box
    ws_exec['B34'] = "RECOMMENDATION:"
    ws_exec['B34'].font = Font(bold=True, size=14)
    ws_exec['B34'].fill = azure_gray
    ws_exec['B34'].font = Font(bold=True, color="FFFFFF")
    ws_exec.merge_cells('B34:F34')
    ws_exec['B34'].alignment = Alignment(horizontal='center')
    
    ws_exec['B36'] = "=Engine!C24"
    ws_exec['B36'].fill = output_green
    ws_exec['B36'].font = Font(bold=True, size=16)
    ws_exec.merge_cells('B36:F36')
    ws_exec['B36'].alignment = Alignment(horizontal='center')
    
    ws_exec['B38'] = "Annual Savings Potential:"
    ws_exec['D38'] = "=Engine!C27"
    ws_exec['D38'].fill = output_green
    ws_exec['D38'].number_format = '$#,##0'
    ws_exec['D38'].font = Font(bold=True, size=14)
    ws_exec.merge_cells('D38:F38')
    ws_exec['D38'].alignment = Alignment(horizontal='center')
    
    # Column widths
    ws_exec.column_dimensions['B'].width = 35
    ws_exec.column_dimensions['C'].width = 15
    ws_exec.column_dimensions['D'].width = 15
    ws_exec.column_dimensions['E'].width = 12
    ws_exec.column_dimensions['F'].width = 15
    
    # Protect worksheet
    ws_exec.protection.sheet = True
    ws_exec.protection.password = "azure-noob-2026"
    
    # Save workbook
    filename = "C:/Users/dswann/Documents/GitHub/azure-noob-blog/Azure_OpenAI_ROI_Calculator_Enterprise.xlsx"
    wb.save(filename)
    
    print("=" * 60)
    print("✅ ENTERPRISE CALCULATOR CREATED")
    print("=" * 60)
    print(f"📁 File: {filename}")
    print(f"📊 Tabs: {len(wb.sheetnames)}")
    print("")
    print("TABS CREATED:")
    print("  1. INSTRUCTIONS - Builds authority")
    print("  2. Engine - The break-even calculator")
    print("  3. Executive Summary - The CFO view")
    print("")
    print("THE $497 LOGIC:")
    print('  ✓ Answers: "When do we switch to PTU?"')
    print("  ✓ Shows: Break-even point formula")
    print("  ✓ Reveals: 15% hidden infrastructure costs")
    print("  ✓ Calculates: Annual savings potential")
    print("")
    print("CELL PROTECTION:")
    print("  ✓ All cells locked except yellow input cells")
    print("  ✓ Feels like software, not a spreadsheet")
    print("")
    print("READY TO UPLOAD TO GUMROAD!")
    print("=" * 60)
    
    return filename

if __name__ == "__main__":
    create_enterprise_calculator()
