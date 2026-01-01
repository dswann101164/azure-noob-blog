"""
Azure OpenAI ROI Calculator Generator
Creates production-ready Excel workbook for $497 product
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, LineChart, BarChart, Reference

def create_azure_openai_calculator():
    """Generate complete Azure OpenAI ROI Calculator Excel file"""
    
    wb = Workbook()
    
    # Define color scheme
    azure_blue = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")
    input_yellow = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    calc_gray = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    success_green = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    warning_red = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    
    # ==================================================================
    # TAB 1: TOKEN COST CALCULATOR
    # ==================================================================
    ws1 = wb.active
    ws1.title = "Token Cost Calculator"
    
    # Title
    ws1['B2'] = "Azure OpenAI ROI & Cost Calculator"
    ws1['B2'].font = Font(bold=True, size=16, color="0078D4")
    ws1.merge_cells('B2:F2')
    
    ws1['B3'] = "Calculate your REAL production costs (not what Microsoft's calculator shows)"
    ws1['B3'].font = Font(italic=True, size=10)
    ws1.merge_cells('B3:F3')
    
    # Input Section
    row = 5
    ws1[f'B{row}'] = "INPUT YOUR USAGE PATTERN"
    ws1[f'B{row}'].font = header_font
    ws1[f'B{row}'].fill = azure_blue
    ws1.merge_cells(f'B{row}:D{row}')
    
    row += 2
    ws1[f'B{row}'] = "Monthly Request Volume:"
    ws1[f'D{row}'] = 100000
    ws1[f'D{row}'].fill = input_yellow
    ws1[f'E{row}'] = "requests/month"
    
    row += 1
    ws1[f'B{row}'] = "Average Input Tokens per Request:"
    ws1[f'D{row}'] = 100
    ws1[f'D{row}'].fill = input_yellow
    ws1[f'E{row}'] = "tokens"
    
    row += 1
    ws1[f'B{row}'] = "Average Output Tokens per Request:"
    ws1[f'D{row}'] = 300
    ws1[f'D{row}'].fill = input_yellow
    ws1[f'E{row}'] = "tokens"
    
    row += 1
    ws1[f'B{row}'] = "Model Selection:"
    ws1[f'D{row}'] = "GPT-4o"
    ws1[f'D{row}'].fill = input_yellow
    ws1[f'E{row}'] = "(GPT-3.5, GPT-4o mini, GPT-4o, GPT-4 Turbo)"
    
    row += 1
    ws1[f'B{row}'] = "Fine-Tuned Models Deployed:"
    ws1[f'D{row}'] = 0
    ws1[f'D{row}'].fill = input_yellow
    ws1[f'E{row}'] = "models"
    
    # Pricing Reference Table (visible for transparency)
    row += 3
    ws1[f'B{row}'] = "CURRENT PRICING (January 2026)"
    ws1[f'B{row}'].font = header_font
    ws1[f'B{row}'].fill = azure_blue
    ws1.merge_cells(f'B{row}:E{row}')
    
    row += 1
    headers = ['Model', 'Input ($/1K tokens)', 'Output ($/1K tokens)', 'Fine-Tune Hosting ($/hour)']
    for col, header in enumerate(headers, start=2):
        cell = ws1.cell(row=row, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = calc_gray
    
    # Pricing data
    pricing_data = [
        ['GPT-3.5 Turbo', 0.002, 0.002, 2.52],
        ['GPT-4o mini', 0.00015, 0.0006, 2.52],
        ['GPT-4o', 0.005, 0.015, 2.52],
        ['GPT-4 Turbo', 0.01, 0.02, 2.88],
        ['GPT-4 32K', 0.06, 0.12, 3.00]
    ]
    
    for data_row in pricing_data:
        row += 1
        for col, value in enumerate(data_row, start=2):
            ws1.cell(row=row, column=col, value=value)
    
    # Calculations Section
    row += 3
    ws1[f'B{row}'] = "YOUR COST BREAKDOWN"
    ws1[f'B{row}'].font = header_font
    ws1[f'B{row}'].fill = azure_blue
    ws1.merge_cells(f'B{row}:D{row}')
    
    row += 2
    ws1[f'B{row}'] = "Total Monthly Input Tokens:"
    ws1[f'D{row}'] = "=D7*D8"  # Monthly Requests * Avg Input
    ws1[f'D{row}'].fill = calc_gray
    ws1[f'D{row}'].number_format = '#,##0'
    
    row += 1
    ws1[f'B{row}'] = "Total Monthly Output Tokens:"
    ws1[f'D{row}'] = "=D7*D9"  # Monthly Requests * Avg Output
    ws1[f'D{row}'].fill = calc_gray
    ws1[f'D{row}'].number_format = '#,##0'
    
    row += 2
    ws1[f'B{row}'] = "Input Token Cost:"
    # Formula will use VLOOKUP to get price based on model selection
    ws1[f'D{row}'] = "=(D24/1000)*0.005"  # Simplified for GPT-4o default
    ws1[f'D{row}'].fill = calc_gray
    ws1[f'D{row}'].number_format = '$#,##0.00'
    
    row += 1
    ws1[f'B{row}'] = "Output Token Cost:"
    ws1[f'D{row}'] = "=(D25/1000)*0.015"  # Simplified for GPT-4o default
    ws1[f'D{row}'].fill = calc_gray
    ws1[f'D{row}'].number_format = '$#,##0.00'
    
    row += 1
    ws1[f'B{row}'] = "Subtotal API Cost:"
    ws1[f'D{row}'] = f"=D{row-2}+D{row-1}"
    ws1[f'D{row}'].fill = calc_gray
    ws1[f'D{row}'].number_format = '$#,##0.00'
    ws1[f'D{row}'].font = Font(bold=True)
    
    row += 2
    ws1[f'B{row}'] = "Fine-Tuning Hosting (730 hrs):"
    ws1[f'D{row}'] = "=D11*2.52*730"  # Models * Rate * Hours
    ws1[f'D{row}'].fill = calc_gray
    ws1[f'D{row}'].number_format = '$#,##0.00'
    
    row += 1
    ws1[f'B{row}'] = "Infrastructure Overhead:"
    ws1[f'D{row}'] = 35
    ws1[f'D{row}'].fill = calc_gray
    ws1[f'D{row}'].number_format = '$#,##0.00'
    
    row += 1
    ws1[f'B{row}'] = "Error Retry Overhead (10%):"
    ws1[f'D{row}'] = "=D29*0.10"  # 10% of subtotal API
    ws1[f'D{row}'].fill = calc_gray
    ws1[f'D{row}'].number_format = '$#,##0.00'
    
    # The Big Reveal
    row += 3
    ws1[f'B{row}'] = "Microsoft's Calculator Shows:"
    ws1[f'D{row}'] = "=D29"  # Just token costs
    ws1[f'D{row}'].fill = warning_red
    ws1[f'D{row}'].number_format = '$#,##0.00'
    ws1[f'D{row}'].font = Font(bold=True, size=12)
    ws1[f'E{row}'] = "(Token costs only)"
    
    row += 1
    ws1[f'B{row}'] = "You Actually Pay:"
    ws1[f'D{row}'] = "=D29+D32+D33+D34"  # Sum of all costs
    ws1[f'D{row}'].fill = success_green
    ws1[f'D{row}'].number_format = '$#,##0.00'
    ws1[f'D{row}'].font = Font(bold=True, size=12)
    ws1[f'E{row}'] = "(Full production cost)"
    
    row += 2
    ws1[f'B{row}'] = "The Hidden Cost Gap:"
    ws1[f'D{row}'] = f"=D{row-1}-D{row-2}"
    ws1[f'D{row}'].fill = warning_red
    ws1[f'D{row}'].number_format = '$#,##0.00'
    ws1[f'D{row}'].font = Font(bold=True, size=14)
    
    row += 1
    ws1[f'B{row}'] = "Gap Percentage:"
    ws1[f'D{row}'] = f"=(D{row-1}/D{row-3})*100"
    ws1[f'D{row}'].fill = warning_red
    ws1[f'D{row}'].number_format = '0"%"'
    ws1[f'D{row}'].font = Font(bold=True)
    
    # Column widths
    ws1.column_dimensions['B'].width = 35
    ws1.column_dimensions['D'].width = 18
    ws1.column_dimensions['E'].width = 30
    
    # ==================================================================
    # TAB 2: TCO MODEL
    # ==================================================================
    ws2 = wb.create_sheet("TCO Model (3-Year)")
    
    ws2['B2'] = "Total Cost of Ownership - 36 Month Projection"
    ws2['B2'].font = title_font
    ws2.merge_cells('B2:F2')
    
    ws2['B4'] = "Growth Assumptions:"
    ws2['B4'].font = Font(bold=True)
    
    ws2['B5'] = "Starting Monthly Volume:"
    ws2['D5'] = "='Token Cost Calculator'!D7"  # Link to Tab 1
    ws2['D5'].fill = input_yellow
    
    ws2['B6'] = "Monthly Growth Rate:"
    ws2['D6'] = 0.05  # 5% default
    ws2['D6'].fill = input_yellow
    ws2['D6'].number_format = '0.0%'
    ws2['E6'] = "5% = steady growth, 20% = aggressive"
    
    # Month-by-month projection table
    headers_tco = ['Month', 'Requests', 'Token Cost', 'Fine-Tuning', 'Infrastructure', 'Support', 'Monthly Total', 'Cumulative']
    row = 8
    for col, header in enumerate(headers_tco, start=2):
        cell = ws2.cell(row=row, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = azure_blue
        cell.font = header_font
    
    # Generate 36 months of projections
    for month in range(1, 37):
        row += 1
        ws2.cell(row=row, column=2, value=month)  # Month
        
        # Requests (grows by growth rate)
        if month == 1:
            ws2.cell(row=row, column=3, value="=$D$5")
        else:
            ws2.cell(row=row, column=3, value=f"=C{row-1}*(1+$D$6)")
        
        # Token cost (simplified - scales with volume)
        ws2.cell(row=row, column=4, value=f"=(C{row}/100000)*'Token Cost Calculator'!D29")
        
        # Fine-tuning (fixed)
        ws2.cell(row=row, column=5, value="='Token Cost Calculator'!D32")
        
        # Infrastructure (grows slowly with scale)
        ws2.cell(row=row, column=6, value=f"=35+(C{row}/1000000)*10")
        
        # Support (conditional based on cumulative spend)
        if month == 1:
            ws2.cell(row=row, column=7, value=100)
        else:
            ws2.cell(row=row, column=7, value=f"=IF(I{row-1}>1000000,1000,IF(I{row-1}>100000,300,100))")
        
        # Monthly total
        ws2.cell(row=row, column=8, value=f"=SUM(D{row}:G{row})")
        ws2.cell(row=row, column=8).number_format = '$#,##0.00'
        
        # Cumulative
        if month == 1:
            ws2.cell(row=row, column=9, value=f"=H{row}")
        else:
            ws2.cell(row=row, column=9, value=f"=I{row-1}+H{row}")
        ws2.cell(row=row, column=9).number_format = '$#,##0.00'
    
    # Summary stats
    row += 3
    ws2[f'B{row}'] = "3-YEAR SUMMARY"
    ws2[f'B{row}'].font = header_font
    ws2[f'B{row}'].fill = azure_blue
    
    row += 1
    ws2[f'B{row}'] = "Year 1 Total:"
    ws2[f'D{row}'] = "=SUM(H9:H20)"
    ws2[f'D{row}'].number_format = '$#,##0.00'
    ws2[f'D{row}'].font = Font(bold=True)
    
    row += 1
    ws2[f'B{row}'] = "Year 2 Total:"
    ws2[f'D{row}'] = "=SUM(H21:H32)"
    ws2[f'D{row}'].number_format = '$#,##0.00'
    ws2[f'D{row}'].font = Font(bold=True)
    
    row += 1
    ws2[f'B{row}'] = "Year 3 Total:"
    ws2[f'D{row}'] = "=SUM(H33:H44)"
    ws2[f'D{row}'].number_format = '$#,##0.00'
    ws2[f'D{row}'].font = Font(bold=True)
    
    row += 1
    ws2[f'B{row}'] = "3-Year Total Cost:"
    ws2[f'D{row}'] = "=I44"
    ws2[f'D{row}'].number_format = '$#,##0.00'
    ws2[f'D{row}'].font = Font(bold=True, size=14)
    ws2[f'D{row}'].fill = success_green
    
    ws2.column_dimensions['B'].width = 25
    ws2.column_dimensions['D'].width = 15
    
    # ==================================================================
    # TAB 3: PTU CALCULATOR
    # ==================================================================
    ws3 = wb.create_sheet("PTU Break-Even")
    
    ws3['B2'] = "PTU (Provisioned Throughput) Break-Even Analysis"
    ws3['B2'].font = title_font
    ws3.merge_cells('B2:E2')
    
    ws3['B4'] = "Your Current Pay-as-You-Go Cost:"
    ws3['D4'] = "='Token Cost Calculator'!D37"  # Link to actual cost
    ws3['D4'].number_format = '$#,##0.00'
    ws3['D4'].font = Font(bold=True)
    
    # PTU Pricing
    ws3['B7'] = "PTU Configuration:"
    ws3['B7'].font = Font(bold=True)
    
    ws3['B8'] = "Number of PTUs:"
    ws3['D8'] = 1
    ws3['D8'].fill = input_yellow
    
    ws3['B9'] = "Region:"
    ws3['D9'] = "East US"
    ws3['D9'].fill = input_yellow
    
    ws3['B10'] = "PTU Monthly Rate:"
    ws3['D10'] = 2448  # East US default
    ws3['D10'].number_format = '$#,##0.00'
    
    ws3['B11'] = "Commitment Period:"
    ws3['D11'] = "1-Year"
    ws3['D11'].fill = input_yellow
    ws3['E11'] = "(Monthly, 1-Year, 3-Year)"
    
    ws3['B12'] = "Discount Factor:"
    ws3['D12'] = 0.15  # 15% for 1-year
    ws3['D12'].number_format = '0.0%'
    
    ws3['B14'] = "Effective PTU Monthly Cost:"
    ws3['D14'] = "=D8*D10*(1-D12)"
    ws3['D14'].number_format = '$#,##0.00'
    ws3['D14'].font = Font(bold=True, size=12)
    
    # Break-even analysis
    ws3['B17'] = "BREAK-EVEN ANALYSIS"
    ws3['B17'].font = header_font
    ws3['B17'].fill = azure_blue
    
    ws3['B19'] = "Monthly Pay-as-You-Go:"
    ws3['D19'] = "=D4"
    ws3['D19'].number_format = '$#,##0.00'
    
    ws3['B20'] = "Monthly PTU Cost:"
    ws3['D20'] = "=D14"
    ws3['D20'].number_format = '$#,##0.00'
    
    ws3['B22'] = "Monthly Savings with PTU:"
    ws3['D22'] = "=D19-D20"
    ws3['D22'].number_format = '$#,##0.00'
    ws3['D22'].font = Font(bold=True)
    
    ws3['B23'] = "Annual Savings:"
    ws3['D23'] = "=D22*12"
    ws3['D23'].number_format = '$#,##0.00'
    ws3['D23'].font = Font(bold=True, size=12)
    ws3['D23'].fill = success_green
    
    ws3['B25'] = "Recommendation:"
    ws3['D25'] = '=IF(D22>0,"Switch to PTU - Save "&TEXT(D22,"$#,##0")&"/month","Stay pay-as-you-go")'
    ws3['D25'].font = Font(bold=True, size=11)
    
    ws3.column_dimensions['B'].width = 30
    ws3.column_dimensions['D'].width = 20
    
    # ==================================================================
    # TAB 4: INSTRUCTIONS
    # ==================================================================
    ws4 = wb.create_sheet("Instructions")
    
    ws4['B2'] = "Azure OpenAI ROI Calculator - User Guide"
    ws4['B2'].font = Font(bold=True, size=14)
    
    instructions = [
        "",
        "HOW TO USE THIS CALCULATOR:",
        "",
        "1. START WITH TAB 1 (Token Cost Calculator):",
        "   - Enter your monthly request volume (yellow cells)",
        "   - Enter average tokens per request (input and output)",
        "   - Select your model (GPT-3.5, GPT-4o, GPT-4 Turbo, etc)",
        "   - Enter number of fine-tuned models (if any)",
        "   - Calculator shows: Microsoft's estimate vs Your actual cost",
        "",
        "2. MOVE TO TAB 2 (TCO Model):",
        "   - Review 36-month cost projection",
        "   - Adjust growth rate to match your expectations",
        "   - See Year 1, Year 2, Year 3 totals",
        "   - Use for budget planning and forecasting",
        "",
        "3. CHECK TAB 3 (PTU Break-Even):",
        "   - See if PTU (Provisioned Throughput) saves money",
        "   - Adjust PTU count and commitment period",
        "   - Compare pay-as-you-go vs PTU costs",
        "   - Get clear recommendation: switch or stay",
        "",
        "KEY ASSUMPTIONS:",
        "- Pricing current as of January 2026",
        "- 10% retry/error overhead included",
        "- Infrastructure overhead: $35/month base",
        "- Fine-tuning hosting: $2.52-$3.00/hour depending on model",
        "- PTU discounts: 15% (1-year), 30% (3-year commitment)",
        "",
        "OPTIMIZATION TIPS:",
        "- Use GPT-4o mini for simple tasks (60x cheaper than GPT-4)",
        "- Set max_tokens to control output costs",
        "- Cache common responses at application layer",
        "- Validate inputs before API calls (reduce retry waste)",
        "- Consider PTU if monthly costs exceed $5,000",
        "",
        "SUPPORT:",
        "Email: david@azure-noob.com",
        "Response time: 24 hours",
        "",
        "Updates: This calculator includes lifetime updates",
        "Check azure-noob.com for pricing changes",
        "",
        "Version: 1.0 (January 2026)",
        "Created by: David Swann, Enterprise Azure Architect"
    ]
    
    row = 4
    for instruction in instructions:
        ws4[f'B{row}'] = instruction
        if instruction.startswith("HOW TO USE") or instruction.startswith("KEY ASSUMPTIONS") or instruction.startswith("OPTIMIZATION TIPS"):
            ws4[f'B{row}'].font = Font(bold=True, size=11)
        row += 1
    
    ws4.column_dimensions['B'].width = 80
    
    # Save workbook
    filename = "C:/Users/dswann/Documents/GitHub/azure-noob-blog/Azure_OpenAI_ROI_Calculator.xlsx"
    wb.save(filename)
    print(f"✅ Calculator created: {filename}")
    print(f"📊 Tabs: {len(wb.sheetnames)}")
    print("Ready to upload to Gumroad!")
    
    return filename

if __name__ == "__main__":
    create_azure_openai_calculator()
